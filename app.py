# app.py
# Streamlit app para consultar y filtrar "PONDERACIONES ASIGNATURAS MOLINA.xlsx"
# Ejecuta: streamlit run app.py

import re
from collections import Counter
import pandas as pd
import openpyxl
import streamlit as st

EXCEL_PATH = "PONDERACIONES ASIGNATURAS MOLINA.xlsx"

# --- utilidades ---
def cell_fill_hex(cell):
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    c = fill.fgColor
    if c is None:
        return None
    if c.type == "rgb" and c.rgb:
        return c.rgb
    return None

@st.cache_data(show_spinner=False)
def load_data(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # asignaturas (fila 1, columnas C..)
    subjects = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            break
        subjects.append(str(v).strip())

    # detectar códigos de universidad frecuentes dentro de paréntesis (URJC, UCM, etc.)
    deg_cells = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
    paren_counter = Counter()
    for s in deg_cells:
        for g in re.findall(r"\(([^()]*)\)", str(s)):
            paren_counter[g.strip()] += 1

    uni_candidates = set()
    for g, cnt in paren_counter.items():
        if (
            cnt >= 5
            and len(g) <= 12
            and re.fullmatch(r"[A-Z0-9/ .-]+", g)
            and (not g.isdigit())
            and (not re.search(r"OPCI|PRESEN|SEMIP|CONJ|INTER|CENTRO|\*", g, re.I))
        ):
            uni_candidates.add(g)
    # fallback
    uni_candidates |= {"UAM", "UCM", "URJC", "UPM", "UC3M", "UAH"}
    # Mapeo de color -> área (según tu criterio)
    color_to_area = {
        "FF99CC00": "RAMA DE CONOCIMIENTO DE ARTES Y HUMANIDADES",   # verde
        "FFFFCC00": "RAMA DE CONOCIMIENTO DE CIENCIAS",              # amarillo
        "FF800080": "RAMA DE CONOCIMIENTO DE CIENCIAS DE LA SALUD",  # morado
        "FF0070BF": "RAMA DE CONOCIMIENTO DE INGENIERÍA Y ARQUITECTURA",  # azul
        "FFFF0000": "RAMA DE CONOCIMIENTO DE CIENCIAS SOCIALES Y JURÍDICAS",  # rojo
    }


    rows = []
    current_color = None
    current_area = "(sin área)"

    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, 2).value
        if full is None:
            continue

        # marca de área: color en col A (solo aparece al empezar un bloque)
        fill = cell_fill_hex(ws.cell(r, 1))
        if fill:
            current_color = fill
            current_area = color_to_area.get(fill, f"Área {fill}")

        s = str(full).strip()
        groups = [g.strip() for g in re.findall(r"\(([^()]*)\)", s)]

        # universidad: primer paréntesis que sea un código conocido o contenga uno
        uni = None
        for g in groups:
            if g in uni_candidates:
                uni = g
                break
        if uni is None:
            for g in groups:
                for cand in uni_candidates:
                    if cand in g:
                        uni = cand
                        break
                if uni:
                    break

        # grado base: texto sin paréntesis
        grado = re.sub(r"\s*\([^()]*\)\s*", " ", s).strip()
        grado = re.sub(r"\s+", " ", grado)

        coeffs = {}
        for i, subj in enumerate(subjects, start=3):
            v = ws.cell(r, i).value
            if v is None or v == "":
                coeffs[subj] = None
            else:
                try:
                    coeffs[subj] = float(str(v).replace(",", "."))
                except Exception:
                    coeffs[subj] = None

        row = {
            "area": current_area,
            "area_color": current_color,
            "grado": grado,
            "universidad": uni,
            "titulo_excel": s,
        }
        row.update(coeffs)
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, subjects, color_to_area

def top4_subjects(row: pd.Series, selected_subjects, only_02: bool):
    items = []
    for subj in selected_subjects:
        v = row.get(subj)
        if pd.isna(v):
            continue
        if only_02 and float(v) != 0.2:
            continue
        items.append((subj, float(v)))
    # ordenar: 0.2 antes que 0.1, luego alfabético
    items.sort(key=lambda x: (-x[1], x[0]))
    return items[:4]

# --- UI ---
st.set_page_config(page_title="Ponderaciones (consulta)", layout="wide")
st.title("Consulta de ponderaciones (Excel)")

df, SUBJECTS, COLOR_TO_AREA = load_data(EXCEL_PATH)

with st.sidebar:
    st.header("Filtros")
    areas = sorted(df["area"].dropna().unique().tolist())
    selected_area = st.selectbox("Área (por color)", areas, index=0)

    q = st.text_input("Buscar grados (palabra/frase)", value="")

    df_area = df[df["area"] == selected_area].copy()
    if q.strip():
        qnorm = q.strip().lower()
        df_area = df_area[df_area["grado"].str.lower().str.contains(qnorm, na=False)]

    grados = sorted(df_area["grado"].dropna().unique().tolist())
    selected_grados = st.multiselect("Selecciona uno o varios grados", grados)

    only_02 = st.checkbox("Solo ponderación 0,2", value=False)

    # asignaturas disponibles (unión dentro de los grados seleccionados si hay selección)
    if selected_grados:
        df_g = df_area[df_area["grado"].isin(selected_grados)]
    else:
        df_g = df_area

    # candidatas: asignaturas con alguna ponderación (0.1/0.2)
    candidate_subjects = []
    for s in SUBJECTS:
        col = df_g[s]
        if only_02:
            ok = (col == 0.2).any()
        else:
            ok = col.notna().any()
        if ok:
            candidate_subjects.append(s)

    st.caption("Asignaturas (solo aparecen si ponderan en lo seleccionado)")
    # botones rápidos
    col1, col2 = st.columns(2)
    if col1.button("Seleccionar todas"):
        st.session_state["sel_subjects"] = candidate_subjects
    if col2.button("Solo las de 0,2"):
        only02_subjects = []
        for s in SUBJECTS:
            if (df_g[s] == 0.2).any():
                only02_subjects.append(s)
        st.session_state["sel_subjects"] = only02_subjects
        st.session_state["only_02"] = True  # no fuerza checkbox, pero ayuda

        # Evitar error cuando cambian área/grados: el default debe ser subset de options
    prev = st.session_state.get("sel_subjects", [])
    safe_default = [s for s in prev if s in candidate_subjects]
    if not safe_default:
        safe_default = candidate_subjects

    selected_subjects = st.multiselect(
        "Asignaturas a considerar",
        options=candidate_subjects,
        default=safe_default,
        key="sel_subjects",
    )

st.subheader("Resultado")

if not selected_grados:
    st.info("Selecciona al menos un **grado** para ver universidades y asignaturas que ponderan.")
    st.stop()

df_sel = df[(df["area"] == selected_area) & (df["grado"].isin(selected_grados))].copy()

# construir salida por universidad (para cada grado)

out_rows = []
for grado in selected_grados:
    df_g = df_sel[df_sel["grado"] == grado].copy()
    if df_g.empty:
        continue
    for _, row in df_g.iterrows():
        top = top4_subjects(row, selected_subjects, only_02)
        
        result_row = {
            "Grado": grado,
            "Universidad": row.get("universidad") or "(sin detectar)",
        }
        
        # Crear columnas separadas para mayor claridad
        for i in range(4):
            if i < len(top):
                result_row[f"Asignatura {i+1}"] = top[i][0]
                result_row[f"Ponderación {i+1}"] = top[i][1]
            else:
                result_row[f"Asignatura {i+1}"] = ""
                result_row[f"Ponderación {i+1}"] = ""
        
        out_rows.append(result_row)

out = pd.DataFrame(out_rows)

if not out.empty:
    out = out.sort_values(["Grado", "Universidad"]).reset_index(drop=True)

st.dataframe(out, use_container_width=True, hide_index=True)


with st.expander("Ver filas originales (título exacto en Excel)"):
    cols = ["area", "grado", "universidad", "titulo_excel"] + SUBJECTS
    st.dataframe(df_sel[cols], use_container_width=True, hide_index=True)
